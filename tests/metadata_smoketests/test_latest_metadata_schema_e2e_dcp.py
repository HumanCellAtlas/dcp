import json
import os
import random
import re
import unittest

from faker import Faker
from ingest.template.schema_template import SchemaTemplate
from ingest.template.vanilla_spreadsheet_builder import VanillaSpreadsheetBuilder
from openpyxl import load_workbook

from ..dataset_fixture import DatasetFixture
from ..dataset_runner import DatasetRunner


class TestLatestMetadataSchemaE2EDcp(unittest.TestCase):
    """ This system test generates a spreadsheet based on the latest metadata schemas at the given deployment (i.e.
    dev, integration, or staging) and runs using a sample SmartSeq2 or 10x dataset."""

    def setUp(self):
        self.faker = Faker()
        self.fixture_name = "latest_metadata"
        self.dataset_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), "..", "fixtures/datasets/",
                                         self.fixture_name)

    def test_latest_metadata_smart_seq_2(self):
        spreadsheet_location = self._setup_metadata_spreadsheet_file_location()
        data_files_location = "s3://org-humancellatlas-dcp-test-data/smart-seq2-one-bundle/"
        number_of_bundles = 1

        self.generate_metadata_spreadsheet(spreadsheet_location, number_of_bundles)
        self._setup_readme_for_dataset_fixture(spreadsheet_location=spreadsheet_location,
                                               data_files_location=data_files_location,
                                               expected_bundle_count=number_of_bundles)

        self._run(fixture_name=self.fixture_name)

    def test_latest_metadata_10x(self):
        spreadsheet_location = self._setup_metadata_spreadsheet_file_location()
        data_files_location = "s3://org-humancellatlas-dcp-test-data/10x/"
        number_of_bundles = 1

        self.generate_metadata_spreadsheet(spreadsheet_location, number_of_bundles)
        self._setup_readme_for_dataset_fixture(spreadsheet_location=spreadsheet_location,
                                               data_files_location=data_files_location,
                                               expected_bundle_count=number_of_bundles)

        self._run(fixture_name=self.fixture_name)

    def generate_metadata_spreadsheet(self, metadata_absolute_file_path, number_of_bundles):
        """ Generate bogus metadata based on the latest metadata schemas and store the fixture. """

        # Create a SchemaTemplate object that encapsulates information about the latest set of metadata schemas.
        ingest_api_url = "http://api.ingest." + os.environ['CI_COMMIT_REF_NAME'] + ".data.humancellatlas.org"
        property_migrations_url = "https://schema." + os.environ[
            'CI_COMMIT_REF_NAME'] + ".data.humancellatlas.org/property_migrations"
        metadata_schema_template = SchemaTemplate(ingest_api_url=ingest_api_url, migrations_url=property_migrations_url)

        # Given the SchemaTemplate that represents all schemas, generate a spreadsheet
        spreadsheet_builder = VanillaSpreadsheetBuilder(output_file=metadata_absolute_file_path, hide_row=False)
        spreadsheet_builder.generate_spreadsheet(schema_template=metadata_schema_template)
        spreadsheet_builder.save_spreadsheet()

        # Generate as many bogus lines of data in the spreadsheet as there is number of bundles.
        self._populate_spreadsheet_with_bogus_data(metadata_absolute_file_path, metadata_schema_template,
                                                   number_of_bundles)

    def _populate_spreadsheet_with_bogus_data(self, spreadsheet_file_path, schema_template, lines_of_bogus_data):
        """ Generate the specified number of bogus lines of metadata and save it in the spreadsheet. The number of
        lines of bogus data in the Sequence tab is equal to the number of bundles that will be generated. """

        spreadsheet = load_workbook(spreadsheet_file_path)

        for tab in spreadsheet.sheetnames:
            worksheet = spreadsheet[tab]

            for property_column_number in range(1, worksheet.max_column + 1):
                description_containing_potential_example = worksheet.cell(row=3, column=property_column_number).value

                # If there is an example that is suggested in the description of the property, just use that.
                # Otherwise, generate an appropriate value based on the type information stored in the SchemaTemplate
                # object that is passed in.
                example_matches = re.search(r'For example: (.*)',
                                            description_containing_potential_example if
                                            description_containing_potential_example else "")
                if example_matches:
                    bogus_data = example_matches.group(1)
                    for row_of_bogus_data in range(6, 6 + lines_of_bogus_data):
                        worksheet.cell(row=row_of_bogus_data, column=property_column_number, value=bogus_data)
                else:
                    fully_qualified_metadata_property = worksheet.cell(row=4, column=property_column_number).value
                    property_attributes = schema_template.lookup_property_attributes_in_metadata(
                        fully_qualified_metadata_property)
                    property_value_type = property_attributes.get('value_type')
                    if property_value_type == "string" and "description" in fully_qualified_metadata_property:
                        bogus_data = self.faker.sentence()
                    elif property_value_type == "string":
                        bogus_data = self.faker.text().split()[0]
                    elif property_value_type == "integer" or property_value_type == "number":
                        bogus_data = random.randint(1, 1000000)
                    elif "microscope" in fully_qualified_metadata_property:
                        bogus_data = "generic confocal"
                    else:
                        print(
                            f"Currently cannot handle generating data of type {property_value_type} for property " \
                                f"{fully_qualified_metadata_property}")

                    for row_of_bogus_data in range(6, 6 + lines_of_bogus_data):
                        worksheet.cell(row=row_of_bogus_data, column=property_column_number, value=bogus_data)

        spreadsheet.save(filename=spreadsheet_file_path)

    def _setup_metadata_spreadsheet_file_location(self):
        """ Creates a new xlsx file that will contain the generated spreadsheet by the test if it doesn't exist. This
        step is necessary because a package that the spreadsheet generator uses (xlsxwriter) expects the file to
        already exist and will throw an error otherwise. """

        fixture_name = "latest_metadata.xlsx"
        full_fixture_path = os.path.join(self.dataset_path, fixture_name)

        # If the file doesn't exist yet, create it because the spreadsheet generator doesn't automatically create
        # spreadsheet files that don't exist.
        if not os.path.exists(full_fixture_path):
            new_spreadsheet_file = open(full_fixture_path, "w+")
            new_spreadsheet_file.close()

        return full_fixture_path

    def _setup_readme_for_dataset_fixture(self, spreadsheet_location, data_files_location, expected_bundle_count):
        """ Write a README file in the test's directory pointing to the local spreadsheet location as well as a
        generic data file location based on the test. Also specify the number of bundles that are expected to be
        generated by the test. """

        full_fixture_readme_path = os.path.join(self.dataset_path, "README.json")

        readme_contents = {"data_files_location": data_files_location, "spreadsheet_location": spreadsheet_location,
                           "expected_bundle_count": expected_bundle_count}

        with open(full_fixture_readme_path, "w+") as readme_file:
            json.dump(readme_contents, readme_file)

    def _run(self, fixture_name, export_bundles=True):
        """ Run the test given the fixture name which contains all pointer to data and metadata files and specify
        whether to export the bundles for downstream analysis. """

        print("")
        dataset = DatasetFixture(fixture_name, deployment=os.environ['CI_COMMIT_REF_NAME'])
        runner = DatasetRunner(deployment=os.environ['CI_COMMIT_REF_NAME'], export_bundles=export_bundles)
        runner.run(dataset_fixture=dataset)
